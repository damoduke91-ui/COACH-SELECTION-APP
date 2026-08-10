"""
AFL Fantasy Stats Fetcher — Scheduled Background Runner
=======================================================
Headless version for Windows Task Scheduler.

This script:
- detects the current AFL round from FootyWire
- detects completed match IDs from FootyWire Stats links
- fetches player stats for completed matches
- builds the XLSX workbook
- saves to AFL_Stats_Latest.xlsx
- saves a timestamped archive in output/
- saves/overwrites the current round file in Round X/

REQUIREMENTS:
    pip install requests beautifulsoup4 openpyxl

USAGE:
    python afl_stats_scheduled_runner.py
"""

import sys
import re
import json
import csv
import datetime
import io
import html as html_lib
import os
import time
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("Missing package: run  pip install requests beautifulsoup4 openpyxl")

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("Missing package: run  pip install beautifulsoup4")

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing package: run  pip install openpyxl")


# ── Project paths ─────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LATEST_XLSX_PATH = SCRIPT_DIR / "AFL_Stats_Latest.xlsx"
LATEST_CSV_PATH = SCRIPT_DIR / "supabase_import_latest.csv"
LOG_PATH = SCRIPT_DIR / "afl_stats_scheduler.log"
STATE_PATH = SCRIPT_DIR / "afl_stats_scheduler_state.json"
EXPECTED_MATCHES_PER_ROUND = 9


def configured_recovery_target() -> tuple[int | None, list[str]]:
    """Return the explicitly locked recovery round and match IDs, if configured."""
    forced_round_text = os.environ.get("AFL_STATS_FORCED_ROUND", "").strip()
    match_ids_text = os.environ.get("AFL_STATS_MATCH_IDS", "").strip()

    if not forced_round_text and not match_ids_text:
        return None, []
    if not forced_round_text or not match_ids_text:
        raise ValueError(
            "Locked recovery requires both AFL_STATS_FORCED_ROUND and AFL_STATS_MATCH_IDS."
        )
    if not forced_round_text.isdigit():
        raise ValueError("AFL_STATS_FORCED_ROUND must be a whole number from 1 to 24.")

    forced_round = int(forced_round_text)
    if forced_round < 1 or forced_round > 24:
        raise ValueError("AFL_STATS_FORCED_ROUND must be a whole number from 1 to 24.")

    supplied_ids = [value.strip() for value in match_ids_text.split(",") if value.strip()]
    if any(not value.isdigit() for value in supplied_ids):
        raise ValueError("Every AFL_STATS_MATCH_IDS value must be a numeric FootyWire match ID.")

    match_ids = sorted(set(supplied_ids), key=int)
    if len(supplied_ids) != EXPECTED_MATCHES_PER_ROUND or len(match_ids) != EXPECTED_MATCHES_PER_ROUND:
        raise ValueError(
            f"Locked recovery requires exactly {EXPECTED_MATCHES_PER_ROUND} unique FootyWire match IDs."
        )

    return forced_round, match_ids

ROUND_FOLDERS = {}
for round_num in range(0, 25):
    round_dir = SCRIPT_DIR / f"Round {round_num}"
    round_dir.mkdir(parents=True, exist_ok=True)
    ROUND_FOLDERS[round_num] = round_dir


# ── Footywire scraping ────────────────────────────────────────────────────────

BASE_URL = "https://www.footywire.com/afl/footy/ft_match_statistics?mid={mid}"
HOME_URL = "https://www.footywire.com/"
SCOREBOARD_URL = "https://www.footywire.com/afl/footy/live_scoreboard"

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

HTTP_SESSION = requests.Session()
HTTP_SESSION.headers.update(HEADERS_HTTP)
RETRYABLE_HTTP_STATUS_CODES = {429, 500, 502, 503, 504}
HTTP_RETRY_DELAYS_SECONDS = (0, 5, 15)

STAT_COLS = [
    "Player", "K", "HB", "D", "M", "G", "B", "T",
    "HO", "GA", "I50", "CL", "CG", "R50", "FF", "FA", "AF", "SC"
]

COL_ALIASES = {
    "Kicks": "K",
    "Handballs": "HB",
    "Disposals": "D",
    "Marks": "M",
    "Goals": "G",
    "Behinds": "B",
    "Tackles": "T",
    "Hit Outs": "HO",
    "HitOuts": "HO",
}

TEAM_ORDER = [
    ("Adelaide", "ADE"),
    ("Brisbane", "BRI"),
    ("Carlton", "CAR"),
    ("Collingwood", "COL"),
    ("Essendon", "ESS"),
    ("Fremantle", "FRE"),
    ("Geelong", "GEE"),
    ("Gold Coast", "GCS"),
    ("GWS", "GWS"),
    ("Hawthorn", "HAW"),
    ("Melbourne", "MEL"),
    ("North Melbourne", "NM"),
    ("Port Adelaide", "PTA"),
    ("Richmond", "RIC"),
    ("St Kilda", "STK"),
    ("Sydney", "SYD"),
    ("West Coast", "WCE"),
    ("Western Bulldogs", "WBU"),
]

TEAM_ALIASES = {
    "adelaide": "Adelaide",
    "adelaide crows": "Adelaide",
    "brisbane": "Brisbane",
    "brisbane lions": "Brisbane",
    "carlton": "Carlton",
    "collingwood": "Collingwood",
    "essendon": "Essendon",
    "fremantle": "Fremantle",
    "freo": "Fremantle",
    "geelong": "Geelong",
    "geelong cats": "Geelong",
    "gold coast": "Gold Coast",
    "gold coast suns": "Gold Coast",
    "gws": "GWS",
    "gws giants": "GWS",
    "greater western sydney": "GWS",
    "greater western sydney giants": "GWS",
    "hawthorn": "Hawthorn",
    "melbourne": "Melbourne",
    "north melbourne": "North Melbourne",
    "north melbourne kangaroos": "North Melbourne",
    "kangaroos": "North Melbourne",
    "port adelaide": "Port Adelaide",
    "port adelaide power": "Port Adelaide",
    "power": "Port Adelaide",
    "richmond": "Richmond",
    "st kilda": "St Kilda",
    "stkilda": "St Kilda",
    "saints": "St Kilda",
    "sydney": "Sydney",
    "sydney swans": "Sydney",
    "swans": "Sydney",
    "west coast": "West Coast",
    "west coast eagles": "West Coast",
    "eagles": "West Coast",
    "western bulldogs": "Western Bulldogs",
    "bulldogs": "Western Bulldogs",
}

KNOWN_TEAMS = [team_name for team_name, _abbr in TEAM_ORDER]
SORTED_TEAM_ALIASES = sorted(TEAM_ALIASES.items(), key=lambda x: len(x[0]), reverse=True)


# ── Logging ───────────────────────────────────────────────────────────────────

def log(message: str) -> None:
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(line + "\n")




# ── Scheduler state / smart-run helpers ───────────────────────────────────────

def load_scheduler_state() -> dict:
    if not STATE_PATH.exists():
        return {"rounds": {}}

    try:
        with STATE_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {"rounds": {}}
        if not isinstance(data.get("rounds"), dict):
            data["rounds"] = {}
        return data
    except Exception as exc:
        log(f"State file could not be read, continuing with a fresh state — {exc}")
        return {"rounds": {}}


def save_scheduler_state(state: dict) -> None:
    temporary_path = STATE_PATH.with_name(f"{STATE_PATH.name}.{datetime.datetime.now().timestamp()}.tmp")
    try:
        temporary_path.write_text(json.dumps(state, indent=2), encoding="utf-8")
        temporary_path.replace(STATE_PATH)
    except Exception as exc:
        log(f"State file could not be written — {exc}")


def sorted_unique_ids(match_ids: list[str]) -> list[str]:
    out: list[str] = []
    for mid in match_ids:
        mid_str = str(mid).strip()
        if mid_str and mid_str not in out:
            out.append(mid_str)
    return sorted(out, key=lambda x: int(x) if x.isdigit() else x)


def get_round_state(state: dict, round_num: int) -> dict:
    rounds = state.setdefault("rounds", {})
    key = str(round_num)
    round_state = rounds.setdefault(key, {})
    if not isinstance(round_state.get("saved_match_ids"), list):
        round_state["saved_match_ids"] = []
    if not isinstance(round_state.get("complete"), bool):
        round_state["complete"] = False
    return round_state


def notify_new_completed_games(round_num: int, new_ids: list[str]) -> None:
    if not new_ids:
        return

    message = (
        f"New completed AFL match detected for Round {round_num}: {', '.join(new_ids)}"
        if len(new_ids) == 1
        else f"New completed AFL matches detected for Round {round_num}: {', '.join(new_ids)}"
    )
    log(message)

    try:
        import winsound  # type: ignore
        winsound.MessageBeep(winsound.MB_ICONASTERISK)
    except Exception:
        pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalise_team_name(name: str) -> str:
    if not name:
        return ""
    cleaned = re.sub(r"\s+", " ", name).strip().lower()
    return TEAM_ALIASES.get(cleaned, re.sub(r"\s+", " ", name).strip())


def team_candidates_from_text(text: str) -> list[str]:
    if not text:
        return []

    cleaned = re.sub(r"\s+", " ", text).strip()
    lower_text = cleaned.lower()

    norm = normalise_team_name(cleaned)
    if norm in KNOWN_TEAMS:
        return [norm]

    found: list[str] = []
    for alias, canonical in SORTED_TEAM_ALIASES:
        pattern = r"(?<![A-Za-z])" + re.escape(alias) + r"(?![A-Za-z])"
        if re.search(pattern, lower_text):
            if canonical not in found:
                found.append(canonical)

    for team_name in KNOWN_TEAMS:
        pattern = r"(?<![A-Za-z])" + re.escape(team_name.lower()) + r"(?![A-Za-z])"
        if re.search(pattern, lower_text):
            if team_name not in found:
                found.append(team_name)

    return found


def guess_team_from_text(text: str) -> str:
    candidates = team_candidates_from_text(text)
    return candidates[0] if candidates else ""


def extract_match_teams_from_title(title: str) -> list[str]:
    if not title:
        return []

    title_clean = re.sub(r"\s+", " ", title).strip()

    split_patterns = [
        r"\s+v\s+",
        r"\s+vs\s+",
        r"\s+V\s+",
        r"\s+VS\s+",
    ]

    for pattern in split_patterns:
        parts = re.split(pattern, title_clean, maxsplit=1)
        if len(parts) == 2:
            left = re.sub(r"^.*?-\s*", "", parts[0]).strip()
            right = re.sub(r"\s*\|.*$", "", parts[1]).strip()
            right = re.sub(r"\s*-\s*.*$", "", right).strip()

            left_team = guess_team_from_text(left)
            right_team = guess_team_from_text(right)

            teams: list[str] = []
            if left_team:
                teams.append(left_team)
            if right_team and right_team != left_team:
                teams.append(right_team)
            return teams

    return []


def fetch_url(url: str) -> str:
    last_error: Exception | None = None

    for attempt, delay_seconds in enumerate(HTTP_RETRY_DELAYS_SECONDS, start=1):
        if delay_seconds:
            log(
                f"Retrying {url} in {delay_seconds} seconds "
                f"(attempt {attempt}/{len(HTTP_RETRY_DELAYS_SECONDS)})"
            )
            time.sleep(delay_seconds)

        try:
            response = HTTP_SESSION.get(url, timeout=20)
            if response.status_code in RETRYABLE_HTTP_STATUS_CODES:
                response.raise_for_status()
            response.raise_for_status()
            return response.text
        except requests.RequestException as exc:
            last_error = exc
            response_status = getattr(getattr(exc, "response", None), "status_code", None)
            if response_status not in RETRYABLE_HTTP_STATUS_CODES:
                raise

    if last_error is not None:
        raise last_error
    raise RuntimeError(f"No HTTP request was attempted for {url}")


def detect_round_and_match_ids() -> tuple[int | None, list[str]]:
    """
    Detect the latest/current round and the completed match IDs for that round.

    FootyWire can briefly show more than one scoreboard area, or the homepage may
    still expose links from the previous round. The old version stopped at the
    first page that had any round + any Stats links, which could keep selecting
    the previous round.

    This version scans all candidate pages, groups Stats links by the nearest
    preceding "Round X Scoreboard" heading in the HTML, then chooses the highest
    round number found. That means if Round 7 has started and only one game has
    finished, it will choose Round 7 with that one completed match instead of
    falling back to Round 6.
    """
    pages_to_try = [HOME_URL, SCOREBOARD_URL]
    ids_by_round: dict[int, list[str]] = {}
    detected_rounds: set[int] = set()

    round_pattern = re.compile(
        r"\b(?:20\d{2}\s+)?Round\s+(\d+)\s+Scoreboard\b",
        re.IGNORECASE,
    )
    anchor_pattern = re.compile(
        r"<a\b[^>]*href=[\"']([^\"']+)[\"'][^>]*>(.*?)</a>",
        re.IGNORECASE | re.DOTALL,
    )

    for url in pages_to_try:
        try:
            page_html = fetch_url(url)
        except Exception as exc:
            log(f"Auto-detect page failed: {url} — {exc}")
            continue

        round_markers: list[tuple[int, int]] = []
        for match in round_pattern.finditer(page_html):
            try:
                marker_round = int(match.group(1))
            except ValueError:
                continue
            round_markers.append((match.start(), marker_round))
            detected_rounds.add(marker_round)

        page_default_round = max((round_num for _pos, round_num in round_markers), default=None)
        page_ids_by_round: dict[int, list[str]] = {}
        orphan_ids: list[str] = []

        for match in anchor_pattern.finditer(page_html):
            href = html_lib.unescape(match.group(1))
            link_html = match.group(2)
            link_text = BeautifulSoup(link_html, "html.parser").get_text(" ", strip=True)
            link_text = re.sub(r"\s+", " ", link_text).strip().lower()

            if link_text != "stats":
                continue

            mid_match = re.search(r"[?&]mid=(\d+)", href)
            if not mid_match:
                continue

            mid = mid_match.group(1)
            nearest_round = None
            for marker_pos, marker_round in round_markers:
                if marker_pos <= match.start():
                    nearest_round = marker_round
                else:
                    break

            if nearest_round is None:
                orphan_ids.append(mid)
                continue

            page_ids = page_ids_by_round.setdefault(nearest_round, [])
            if mid not in page_ids:
                page_ids.append(mid)

        # If the page only contains one scoreboard heading but the links appear
        # outside/above that heading in the source, use the page round as a safe
        # fallback for those orphan Stats links.
        if page_default_round is not None and orphan_ids:
            page_ids = page_ids_by_round.setdefault(page_default_round, [])
            for mid in orphan_ids:
                if mid not in page_ids:
                    page_ids.append(mid)

        for round_num, ids in sorted(page_ids_by_round.items()):
            global_ids = ids_by_round.setdefault(round_num, [])
            for mid in ids:
                if mid not in global_ids:
                    global_ids.append(mid)

        found_text = ", ".join(
            f"Round {round_num}: {len(ids)} Stats link(s)"
            for round_num, ids in sorted(page_ids_by_round.items())
        ) or "no Stats links"
        rounds_text = ", ".join(f"Round {r}" for r in sorted(detected_rounds)) or "no round headings"
        log(f"Auto-detect scanned {url} — headings: {rounds_text}; {found_text}")

    if ids_by_round:
        chosen_round = max(ids_by_round.keys())
        chosen_ids = ids_by_round[chosen_round][:9]
        log(f"Auto-detect selected Round {chosen_round} with {len(chosen_ids)} completed match ID(s)")
        return chosen_round, chosen_ids

    if detected_rounds:
        chosen_round = max(detected_rounds)
        log(f"Auto-detect selected Round {chosen_round}, but found no completed match IDs yet")
        return chosen_round, []

    return None, []


def fetch_match(mid: str) -> dict:
    url = BASE_URL.format(mid=mid)
    return parse_match(fetch_url(url), mid)


def parse_match(html: str, mid: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")

    page_text = soup.get_text(" ", strip=True)
    page_round_match = re.search(r"\bRound\s+(\d+)\b", page_text, re.IGNORECASE)
    page_round = int(page_round_match.group(1)) if page_round_match else None

    title = f"Match {mid}"
    for sel in [".lnormtitle", "h1", ".lheader", "title"]:
        el = soup.select_one(sel)
        if el:
            t = el.get_text(" ", strip=True)
            if 3 < len(t) < 140:
                title = re.sub(r"\s+", " ", t).strip()
                break

    title_teams = extract_match_teams_from_title(title)
    teams: list[dict] = []

    for tbl in soup.find_all("table"):
        header_row = tbl.find("tr")
        if not header_row:
            continue

        hcells = [c.get_text(strip=True) for c in header_row.find_all(["th", "td"])]
        if "Player" not in hcells or len(hcells) < 8:
            continue

        col_map: dict[int, str] = {}
        for i, h in enumerate(hcells):
            canonical = COL_ALIASES.get(h, h)
            if canonical in STAT_COLS:
                col_map[i] = canonical

        team_name_candidates: list[str] = []

        prev = tbl.find_previous_sibling()
        while prev:
            txt = re.sub(r"\s+", " ", prev.get_text(" ", strip=True)).strip()
            if 2 < len(txt) < 100:
                candidates = team_candidates_from_text(txt)
                if candidates:
                    team_name_candidates.extend(candidates)
                    break
            prev = prev.find_previous_sibling()

        if not team_name_candidates:
            parent = tbl.parent
            for _ in range(3):
                if not parent:
                    break
                txt = re.sub(r"\s+", " ", parent.get_text(" ", strip=True)).strip()
                candidates = team_candidates_from_text(txt)
                if candidates:
                    team_name_candidates.extend(candidates)
                    break
                parent = parent.parent

        team_name = team_name_candidates[0] if team_name_candidates else ""

        players: list[dict] = []
        for tr in tbl.find_all("tr")[1:]:
            cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if not cells or len(cells) < 5:
                continue

            player_name = cells[0] if cells else ""
            if not player_name or player_name in ("Total", "Player", ""):
                continue
            if not re.search(r"[A-Za-z]{2}", player_name):
                continue

            row: dict[str, str] = {}
            for idx, canon in col_map.items():
                row[canon] = cells[idx] if idx < len(cells) else ""

            if not row.get("Player"):
                row["Player"] = cells[0]

            if row.get("Player"):
                players.append(row)

        if players:
            teams.append({"name": team_name, "players": players})

    if not teams:
        raise ValueError(
            f"No player stats found. The match may not have stats yet, or mid={mid} is incorrect."
        )

    if len(title_teams) == 2 and len(teams) == 2:
        teams[0]["name"] = title_teams[0]
        teams[1]["name"] = title_teams[1]
    else:
        if title_teams:
            title_idx = 0
            used = set()
            for team in teams:
                if team["name"]:
                    used.add(team["name"])

            for team in teams:
                if not team["name"]:
                    while title_idx < len(title_teams) and title_teams[title_idx] in used:
                        title_idx += 1
                    if title_idx < len(title_teams):
                        team["name"] = title_teams[title_idx]
                        used.add(title_teams[title_idx])
                        title_idx += 1

    return {"mid": mid, "title": title, "round": page_round, "teams": teams}


# ── Excel building ────────────────────────────────────────────────────────────

C_NAVY = "0A1628"
C_GOLD = "F5C842"
C_WHITE = "FFFFFF"
C_LIGHT = "EEF3FA"
C_TEAM = "1A3060"
C_HEADER = "0F2040"


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def make_font(bold=False, color=C_WHITE, size=11, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)


def thin_border():
    s = Side(style="thin", color="D0D8E8")
    return Border(left=s, right=s, top=s, bottom=s)


def to_number_if_possible(val):
    if val in ("", None):
        return val
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return float(val)
        except (ValueError, TypeError):
            return val


def style_data_cell(cell, col_name, row_fill, left_align=False, bold=False):
    font_data = make_font(bold=False, color=C_NAVY, size=10)
    font_bold = make_font(bold=True, color=C_NAVY, size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    cell.fill = row_fill
    cell.border = thin_border()

    if col_name == "Player" or left_align:
        cell.font = font_bold if bold else font_data
        cell.alignment = left
    elif col_name in ("SC", "AF"):
        cell.font = Font(
            bold=True,
            color="0A5C2A" if col_name == "SC" else C_NAVY,
            size=10,
            name="Arial"
        )
        cell.alignment = center
    else:
        cell.font = font_bold if bold else font_data
        cell.alignment = center


def write_match_sheet(ws, match_data):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    fill_navy = make_fill(C_NAVY)
    fill_header = make_fill(C_HEADER)
    fill_team = make_fill(C_TEAM)
    fill_alt = make_fill(C_LIGHT)
    fill_white = make_fill(C_WHITE)

    font_title = make_font(bold=True, color=C_GOLD, size=14)
    font_header = make_font(bold=True, color=C_GOLD, size=10)
    font_team = make_font(bold=True, color=C_GOLD, size=10)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{get_column_letter(len(STAT_COLS))}1")
    tc = ws["A1"]
    tc.value = f"{match_data['title']}  [mid={match_data['mid']}]"
    tc.fill = fill_navy
    tc.font = font_title
    tc.alignment = left

    ws.row_dimensions[2].height = 22
    for ci, col in enumerate(STAT_COLS, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = center
        cell.border = thin_border()

    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(STAT_COLS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6

    current_row = 3
    for team in match_data["teams"]:
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(f"A{current_row}:{get_column_letter(len(STAT_COLS))}{current_row}")
        tc = ws.cell(row=current_row, column=1, value=team["name"] or "Team")
        tc.fill = fill_team
        tc.font = font_team
        tc.alignment = left
        tc.border = thin_border()
        current_row += 1

        for pi, player in enumerate(team["players"]):
            ws.row_dimensions[current_row].height = 18
            row_fill = fill_alt if pi % 2 == 0 else fill_white

            for ci, col in enumerate(STAT_COLS, 1):
                val = to_number_if_possible(player.get(col, ""))
                cell = ws.cell(row=current_row, column=ci, value=val)
                style_data_cell(cell, col, row_fill)

            current_row += 1

        current_row += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(STAT_COLS))}{current_row}"


def write_summary_sheet(ws, all_matches):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    fill_navy = make_fill(C_NAVY)
    fill_header = make_fill(C_HEADER)
    fill_alt = make_fill(C_LIGHT)
    fill_white = make_fill(C_WHITE)

    font_title = make_font(bold=True, color=C_GOLD, size=14)
    font_header = make_font(bold=True, color=C_GOLD, size=10)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    all_cols = ["Match", "Team"] + STAT_COLS

    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{get_column_letter(len(all_cols))}1")
    tc = ws["A1"]
    tc.value = f"All Matches Combined — {datetime.date.today().strftime('%d %b %Y')}"
    tc.fill = fill_navy
    tc.font = font_title
    tc.alignment = left

    ws.row_dimensions[2].height = 22
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = center
        cell.border = thin_border()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    for ci in range(4, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6

    current_row = 3
    row_counter = 0

    for match in all_matches:
        for team in match["teams"]:
            for player in team["players"]:
                row_fill = fill_alt if row_counter % 2 == 0 else fill_white
                row_data = [match["title"], team["name"]] + [player.get(c, "") for c in STAT_COLS]

                for ci, val in enumerate(row_data, 1):
                    col_name = all_cols[ci - 1]
                    val = to_number_if_possible(val)
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    style_data_cell(
                        cell,
                        col_name,
                        row_fill,
                        left_align=(col_name in ("Match", "Team")),
                        bold=(col_name == "Player")
                    )

                current_row += 1
                row_counter += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(all_cols))}{current_row}"


def collect_teams_for_all_teams_sheet(all_matches):
    teams_map = {team_name: [] for team_name, _abbr in TEAM_ORDER}

    for match in all_matches:
        for team in match.get("teams", []):
            team_name = normalise_team_name(team.get("name", ""))
            players = team.get("players", [])

            if team_name in teams_map and players:
                teams_map[team_name] = players

    return teams_map


def write_all_teams_sheet(ws, all_matches):
    ws.sheet_view.showGridLines = False

    fill_header = make_fill(C_HEADER)
    fill_team = make_fill(C_TEAM)
    fill_alt = make_fill(C_LIGHT)
    fill_white = make_fill(C_WHITE)

    font_team = make_font(bold=True, color=C_GOLD, size=12)
    font_header = make_font(bold=True, color=C_GOLD, size=10)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    teams_map = collect_teams_for_all_teams_sheet(all_matches)

    ws.column_dimensions["A"].width = 26
    for ci in range(2, len(STAT_COLS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6

    for team_index, (team_name, team_abbr) in enumerate(TEAM_ORDER):
        start_row = 1 + (team_index * 30)
        header_row = start_row + 1
        data_row_start = start_row + 2

        ws.row_dimensions[start_row].height = 22
        ws.merge_cells(f"A{start_row}:{get_column_letter(len(STAT_COLS))}{start_row}")
        title_cell = ws.cell(row=start_row, column=1, value=f"{team_name} ({team_abbr})")
        title_cell.fill = fill_team
        title_cell.font = font_team
        title_cell.alignment = left
        title_cell.border = thin_border()

        for ci, col in enumerate(STAT_COLS, 1):
            cell = ws.cell(row=header_row, column=ci, value=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = center
            cell.border = thin_border()

        players = teams_map.get(team_name, [])
        for pi, player in enumerate(players):
            row_num = data_row_start + pi
            row_fill = fill_alt if pi % 2 == 0 else fill_white

            for ci, col in enumerate(STAT_COLS, 1):
                val = to_number_if_possible(player.get(col, ""))
                cell = ws.cell(row=row_num, column=ci, value=val)
                style_data_cell(cell, col, row_fill)

        block_end = start_row + 29
        used_end = data_row_start + len(players) - 1
        for row_num in range(max(data_row_start, used_end + 1), block_end + 1):
            for ci in range(1, len(STAT_COLS) + 1):
                cell = ws.cell(row=row_num, column=ci, value=None)
                cell.fill = fill_white
                cell.border = thin_border()

    ws.freeze_panes = "A2"


def build_xlsx(all_matches) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("All Matches")
    write_summary_sheet(ws_summary, all_matches)

    ws_all_teams = wb.create_sheet("All Teams", 1)
    write_all_teams_sheet(ws_all_teams, all_matches)

    for match in all_matches:
        safe_name = re.sub(r'[\\/*?:\[\]]', "", match["title"])[:28] or f"Match {match['mid']}"
        ws = wb.create_sheet(safe_name)
        write_match_sheet(ws, match)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Supabase CSV export ──────────────────────────────────────────────────────

SUPABASE_ENVIRONMENT = os.environ.get("AFL_STATS_ENVIRONMENT", "preview").strip().lower()
if SUPABASE_ENVIRONMENT not in {"preview", "production"}:
    sys.exit("AFL_STATS_ENVIRONMENT must be either preview or production.")

SUPABASE_CSV_COLS = [
    "environment", "afl_round", "afl_team_name", "afl_team_code", "player_name",
    "k", "hb", "d", "m", "g", "b", "t", "ho", "ga", "i50", "cl", "cg", "r50",
    "ff", "fa", "af", "sc",
]

TEAM_CODE_BY_NAME = {team_name: team_code for team_name, team_code in TEAM_ORDER}

STAT_TO_CSV_FIELD = {
    "K": "k",
    "HB": "hb",
    "D": "d",
    "M": "m",
    "G": "g",
    "B": "b",
    "T": "t",
    "HO": "ho",
    "GA": "ga",
    "I50": "i50",
    "CL": "cl",
    "CG": "cg",
    "R50": "r50",
    "FF": "ff",
    "FA": "fa",
    "AF": "af",
    "SC": "sc",
}


def to_csv_number(value) -> int | float | str:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return value

    cleaned = str(value).replace(",", "").strip()
    if cleaned == "":
        return 0

    try:
        return int(cleaned)
    except ValueError:
        try:
            return float(cleaned)
        except ValueError:
            return cleaned


def safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", str(value or "").strip()).strip("_")
    return cleaned or "UNKNOWN"


def team_code_for_name(team_name: str) -> str:
    return TEAM_CODE_BY_NAME.get(normalise_team_name(team_name), "")


def match_team_codes(match: dict) -> list[str]:
    codes: list[str] = []
    for team in match.get("teams", []):
        code = team_code_for_name(team.get("name", ""))
        if code and code not in codes:
            codes.append(code)
    return codes


def per_game_csv_filename(match: dict, round_num: int) -> str:
    codes = match_team_codes(match)
    if len(codes) >= 2:
        return f"round{round_num}_stats_{safe_filename_part(codes[0])}_{safe_filename_part(codes[1])}.csv"
    if len(codes) == 1:
        return f"round{round_num}_stats_{safe_filename_part(codes[0])}.csv"
    mid = safe_filename_part(match.get("mid", "match"))
    return f"round{round_num}_stats_match_{mid}.csv"


def progressive_csv_filename(round_num: int) -> str:
    return f"supabase_import_round_{round_num}_progressive.csv"


def timestamped_progressive_csv_name(round_num: int | None = None) -> str:
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if round_num is None:
        return f"supabase_import_progressive_{ts}.csv"
    return f"supabase_import_round_{round_num}_progressive_{ts}.csv"


def row_stat_score(row: dict) -> float:
    """
    Used only for de-duplicating FootyWire placeholder rows.
    A genuine zero-stat player is kept when there is no duplicate non-zero row.
    """
    total = 0.0
    for field in STAT_TO_CSV_FIELD.values():
        value = row.get(field, 0)
        if isinstance(value, (int, float)):
            total += abs(float(value))
            continue
        try:
            total += abs(float(str(value).strip() or 0))
        except (TypeError, ValueError):
            pass
    return total


def value_from_player(player: dict, stat_col: str):
    """Read stats from either uppercase scraper keys or lowercase JSON keys."""
    csv_field = STAT_TO_CSV_FIELD[stat_col]
    return player.get(stat_col, player.get(csv_field, 0))


def build_supabase_csv_rows(all_matches: list[dict], round_num: int) -> list[dict]:
    rows: list[dict] = []

    for match in all_matches:
        for team in match.get("teams", []):
            team_name = normalise_team_name(team.get("name", ""))
            team_code = TEAM_CODE_BY_NAME.get(team_name, "")

            for player in team.get("players", []):
                player_name = str(player.get("Player", player.get("player_name", ""))).strip()

                # Remove FootyWire header rows accidentally scraped as players
                if not player_name:
                    continue

                if "match statistics" in player_name.lower():
                    continue

                stat_values = {
                    csv_field: to_csv_number(value_from_player(player, stat_col))
                    for stat_col, csv_field in STAT_TO_CSV_FIELD.items()
                }

                row = {
                    "environment": SUPABASE_ENVIRONMENT,
                    "afl_round": round_num,
                    "afl_team_name": team_name,
                    "afl_team_code": team_code,
                    "player_name": player_name,
                    **stat_values,
                }

                rows.append(row)

    # FootyWire/local-server payloads can include duplicate placeholder rows first,
    # followed by the real player rows. Do NOT remove every all-zero player row,
    # because a real player can occasionally finish with zero stats.
    # Instead, de-duplicate by round/team/player and keep the row with actual stats
    # when one exists. If every duplicate is zero, keep a single zero row.
    best_by_player: dict[tuple, dict] = {}
    order: list[tuple] = []

    for row in rows:
        key = (
            row.get("environment", ""),
            row.get("afl_round", ""),
            # Deliberately ignore team code in de-dupe because
            # FootyWire placeholder rows can incorrectly assign
            # the opponent team code to duplicated zero-stat rows.
            row.get("player_name", ""),
        )

        if key not in best_by_player:
            best_by_player[key] = row
            order.append(key)
            continue

        existing = best_by_player[key]
        if row_stat_score(row) > row_stat_score(existing):
            best_by_player[key] = row

    return [best_by_player[key] for key in order]

def build_supabase_csv_text(all_matches: list[dict], round_num: int) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=SUPABASE_CSV_COLS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(build_supabase_csv_rows(all_matches, round_num))
    return output.getvalue()


def write_csv_text(path: Path, csv_text: str) -> None:
    if path.exists():
        path.unlink()
    path.write_text(csv_text, encoding="utf-8", newline="")


def save_supabase_csv_files(all_matches: list[dict], round_num: int | None = None) -> dict:
    """
    Saves both CSV formats:
      1. One progressive round CSV containing all completed games fetched so far.
      2. One per-game CSV for each match, matching the uploaded template structure.
    """
    if round_num is None:
        raise ValueError("round_num is required to save the Supabase import CSV")

    round_dir = ROUND_FOLDERS[round_num]

    progressive_csv_text = build_supabase_csv_text(all_matches, round_num)
    latest_progressive_path = LATEST_CSV_PATH
    archive_progressive_path = OUTPUT_DIR / timestamped_progressive_csv_name(round_num)
    round_progressive_path = round_dir / progressive_csv_filename(round_num)

    write_csv_text(latest_progressive_path, progressive_csv_text)
    write_csv_text(archive_progressive_path, progressive_csv_text)
    write_csv_text(round_progressive_path, progressive_csv_text)

    per_game_paths: list[Path] = []
    for match in all_matches:
        match_csv_text = build_supabase_csv_text([match], round_num)
        match_path = round_dir / per_game_csv_filename(match, round_num)
        write_csv_text(match_path, match_csv_text)
        per_game_paths.append(match_path)

    return {
        "latest_csv_path": latest_progressive_path,
        "archive_csv_path": archive_progressive_path,
        "round_csv_path": round_progressive_path,
        "per_game_csv_paths": per_game_paths,
    }


def timestamped_xlsx_name() -> str:
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    return f"AFL_Stats_{ts}.xlsx"


def round_filename(round_num: int) -> str:
    return f"AFL_Stats_Round_{round_num}.xlsx"


def save_xlsx_files(xlsx_bytes: bytes, round_num: int | None = None) -> tuple[Path, Path, Path | None]:
    archive_path = OUTPUT_DIR / timestamped_xlsx_name()
    archive_path.write_bytes(xlsx_bytes)

    if LATEST_XLSX_PATH.exists():
        LATEST_XLSX_PATH.unlink()
    LATEST_XLSX_PATH.write_bytes(xlsx_bytes)

    round_path = None
    if round_num is not None:
        round_dir = ROUND_FOLDERS[round_num]
        round_path = round_dir / round_filename(round_num)
        if round_path.exists():
            round_path.unlink()
        round_path.write_bytes(xlsx_bytes)

    return LATEST_XLSX_PATH, archive_path, round_path


# ── Main scheduled run ────────────────────────────────────────────────────────

def run_scheduled_fetch() -> int:
    log("=" * 60)
    log("Scheduled AFL stats run started")

    try:
        forced_round, forced_match_ids = configured_recovery_target()
        recovery_mode = forced_round is not None

        if recovery_mode:
            round_num = forced_round
            match_ids = forced_match_ids
            log(
                f"Locked recovery mode: AFL Round {round_num} with exactly "
                f"{len(match_ids)} explicitly supplied match IDs."
            )
        else:
            round_num, match_ids = detect_round_and_match_ids()
        if round_num is None:
            log("Could not detect the current round number from FootyWire.")
            return 1

        match_ids = sorted_unique_ids(match_ids)
        state = load_scheduler_state()
        round_state = get_round_state(state, round_num)
        previous_ids = sorted_unique_ids(round_state.get("saved_match_ids", []))

        log(f"Detected Round {round_num}")

        if not match_ids:
            log(f"No completed match IDs found for Round {round_num}. Nothing to fetch yet.")
            round_state["last_checked_at"] = datetime.datetime.now().isoformat(timespec="seconds")
            save_scheduler_state(state)
            return 0

        log(f"Detected {len(match_ids)} completed match ID(s): {', '.join(match_ids)}")

        new_ids = [mid for mid in match_ids if mid not in previous_ids]
        notify_new_completed_games(round_num, new_ids)

        round_is_complete = len(match_ids) >= EXPECTED_MATCHES_PER_ROUND

        if not recovery_mode and previous_ids == match_ids and LATEST_XLSX_PATH.exists():
            if round_state.get("complete"):
                log(
                    f"Round {round_num} is already marked complete with "
                    f"{len(match_ids)} match ID(s). No new XLSX needed."
                )
            else:
                log(
                    f"No new completed matches for Round {round_num}; "
                    f"already saved {len(match_ids)} match ID(s). Skipping fetch and save."
                )

            round_state["last_checked_at"] = datetime.datetime.now().isoformat(timespec="seconds")
            save_scheduler_state(state)
            return 0

        all_matches: list[dict] = []
        total_players = 0

        for mid in match_ids:
            try:
                match_data = fetch_match(mid)
                player_count = sum(len(team.get("players", [])) for team in match_data.get("teams", []))

                if recovery_mode:
                    detected_match_round = match_data.get("round")
                    if detected_match_round != round_num:
                        raise ValueError(
                            f"Safety stop: match {mid} identifies as Round "
                            f"{detected_match_round or 'unknown'}, not requested Round {round_num}."
                        )
                    if player_count < 40:
                        raise ValueError(
                            f"Safety stop: match {mid} contains only {player_count} player rows."
                        )

                total_players += player_count
                all_matches.append(match_data)
                log(f"Fetched mid={mid} — {match_data['title']} ({player_count} players)")
            except Exception as exc:
                log(f"Failed mid={mid} — {exc}")

        if recovery_mode and len(all_matches) != EXPECTED_MATCHES_PER_ROUND:
            log(
                "Locked recovery stopped before file generation: "
                f"validated {len(all_matches)} of {EXPECTED_MATCHES_PER_ROUND} required matches."
            )
            return 1

        if not all_matches:
            log("No matches were fetched successfully. XLSX was not saved.")
            return 1

        xlsx_bytes = build_xlsx(all_matches)
        latest_path, archive_path, round_path = save_xlsx_files(xlsx_bytes, round_num)
        csv_paths = save_supabase_csv_files(all_matches, round_num)
        latest_csv_path = csv_paths["latest_csv_path"]
        archive_csv_path = csv_paths["archive_csv_path"]
        round_csv_path = csv_paths["round_csv_path"]
        per_game_csv_paths = csv_paths["per_game_csv_paths"]

        now_iso = datetime.datetime.now().isoformat(timespec="seconds")
        round_state["saved_match_ids"] = match_ids
        round_state["match_count"] = len(match_ids)
        round_state["total_players"] = total_players
        round_state["last_saved_at"] = now_iso
        round_state["last_checked_at"] = now_iso
        round_state["complete"] = round_is_complete
        round_state["latest_path"] = str(latest_path)
        round_state["round_path"] = str(round_path) if round_path else ""
        round_state["latest_csv_path"] = str(latest_csv_path)
        round_state["round_csv_path"] = str(round_csv_path) if round_csv_path else ""
        round_state["per_game_csv_paths"] = [str(path) for path in per_game_csv_paths]
        state["last_run"] = {
            "round": round_num,
            "match_count": len(match_ids),
            "total_players": total_players,
            "saved_at": now_iso,
            "complete": round_is_complete,
        }
        save_scheduler_state(state)

        log(f"Successful fetches: {len(all_matches)}")
        log(f"Total players loaded: {total_players}")
        log(f"Latest XLSX saved to:  {latest_path}")
        log(f"Archive XLSX saved to: {archive_path}")
        log(f"Round XLSX saved to:   {round_path}")
        log(f"Latest progressive CSV saved to:   {latest_csv_path}")
        log(f"Archive progressive CSV saved to:  {archive_csv_path}")
        log(f"Round progressive CSV saved to:    {round_csv_path}")
        for path in per_game_csv_paths:
            log(f"Per-game CSV saved to:             {path}")

        if round_is_complete:
            log(
                f"Round {round_num} is now marked complete "
                f"({len(match_ids)}/{EXPECTED_MATCHES_PER_ROUND} matches)."
            )
        else:
            log(
                f"Round {round_num} is still in progress "
                f"({len(match_ids)}/{EXPECTED_MATCHES_PER_ROUND} matches)."
            )

        log("Scheduled AFL stats run finished successfully")
        return 0

    except Exception as exc:
        log(f"Scheduled run failed: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(run_scheduled_fetch())
